#!/usr/bin/env python3
from __future__ import annotations

# Standard Library
from pathlib import Path
import csv

try:
	from odf import table as odf_table
	from odf import text as odf_text
	from odf.opendocument import load as odf_load
except Exception:
	odf_table = None
	odf_text = None
	odf_load = None

try:
	import openpyxl
except Exception:
	openpyxl = None

try:
	import xlrd
except Exception:
	xlrd = None

# local repo modules
from .base import FileMetadata, FileMetadataPlugin
from .mdls_utils import mdls_field

#============================================


class SpreadsheetPlugin(FileMetadataPlugin):
	"""
	Plugin for spreadsheet and delimited files.
	"""

	name = "spreadsheet"
	supported_suffixes: set[str] = {"xls", "xlsx", "ods", "csv", "tsv"}

	#============================================
	def supports(self, path: Path) -> bool:
		return path.suffix.lower().lstrip(".") in self.supported_suffixes

	#============================================
	def extract_metadata(self, path: Path) -> FileMetadata:
		meta = FileMetadata(path=path, plugin_name=self.name)
		meta.extra["size_bytes"] = path.stat().st_size
		meta.extra["extension"] = path.suffix.lstrip(".")
		title = mdls_field(path, "kMDItemTitle")
		if title:
			meta.title = title
		meta.summary = self._build_summary(path)
		return meta

	#============================================
	def _build_summary(self, path: Path) -> str:
		"""
		Construct a summary using headers/rows or sheet names.
		"""
		ext = path.suffix.lower().lstrip(".")
		if ext in {"csv", "tsv"}:
			preview = self._read_delimited(path, delimiter="\t" if ext == "tsv" else ",")
			if preview:
				return preview
		if ext == "ods":
			preview = self._ods_preview(path)
			if preview:
				return preview
		if ext == "xlsx":
			preview = self._xlsx_preview(path)
			if preview:
				return preview
		if ext == "xls":
			preview = self._xls_preview(path)
			if preview:
				return preview
		return f"Data file {path.name}"

	#============================================
	def _read_delimited(self, path: Path, delimiter: str) -> str | None:
		"""
		Read first few lines for delimited text.
		"""
		try:
			with path.open("r", encoding="utf-8", errors="ignore") as handle:
				reader = csv.reader(handle, delimiter=delimiter)
				rows = []
				for idx, row in enumerate(reader):
					rows.append(row)
					if idx >= 2:
						break
			if not rows:
				return None
			joined_rows = [" | ".join(r) for r in rows]
			head = " || ".join(joined_rows)
			head = " ".join(head.split())
			if not head:
				return None
			return head[:800]
		except Exception:
			return None

	#============================================
	#============================================
	def _xlsx_preview(self, path: Path) -> str | None:
		"""
		Extract sheet names plus first row and first column previews.
		"""
		if not openpyxl:
			return None
		try:
			wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
			sheet_names = list(wb.sheetnames)
			ws = wb[sheet_names[0]] if sheet_names else wb.active
			row1_values: list[str] = []
			for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
				for value in row:
					if value is None:
						continue
					text = str(value).strip()
					if text:
						row1_values.append(text)
					if len(row1_values) >= 8:
						break
			col1_values: list[str] = []
			for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
				value = row[0] if row else None
				if value is None:
					continue
				text = str(value).strip()
				if text:
					col1_values.append(text)
				if len(col1_values) >= 8:
					break
			parts: list[str] = []
			if sheet_names:
				parts.append(f"Sheets: {', '.join(sheet_names[:5])}")
			if row1_values:
				parts.append("Row1: " + " | ".join(row1_values[:8]))
			if col1_values:
				parts.append("Col1: " + " | ".join(col1_values[:8]))
			summary = " || ".join(parts).strip()
			return summary[:800] if summary else None
		except Exception:
			return None

	#============================================
	def _xls_preview(self, path: Path) -> str | None:
		if not xlrd:
			return None
		try:
			book = xlrd.open_workbook(path)
			sheet_names = book.sheet_names()
			sheet = book.sheet_by_index(0)
			row1_values = [
				str(value).strip()
				for value in sheet.row_values(0)[:8]
				if value not in (None, "")
			]
			col1_values = [
				str(value).strip()
				for value in sheet.col_values(0)[:8]
				if value not in (None, "")
			]
			parts: list[str] = []
			if sheet_names:
				parts.append(f"Sheets: {', '.join(sheet_names[:5])}")
			if row1_values:
				parts.append("Row1: " + " | ".join(row1_values[:8]))
			if col1_values:
				parts.append("Col1: " + " | ".join(col1_values[:8]))
			summary = " || ".join(parts).strip()
			return summary[:800] if summary else None
		except Exception:
			return None

	#============================================
	def _ods_preview(self, path: Path) -> str | None:
		if not odf_load or not odf_table or not odf_text:
			return None
		try:
			document = odf_load(str(path))
			sheets = document.spreadsheet.getElementsByType(odf_table.Table)
			if not sheets:
				return None
			names: list[str] = []
			preview_rows: list[str] = []
			col1_values: list[str] = []
			for sheet in sheets:
				name = sheet.getAttribute("name")
				if name:
					names.append(name)
				if preview_rows:
					continue
				for row in sheet.getElementsByType(odf_table.TableRow):
					row_values: list[str] = []
					first_col: str | None = None
					for cell in row.getElementsByType(odf_table.TableCell):
						paras = cell.getElementsByType(odf_text.P)
						text_bits = []
						for para in paras:
							if para.firstChild:
								text_bits.append(str(para.firstChild.data))
						cell_text = " ".join(text_bits).strip()
						if cell_text:
							row_values.append(cell_text)
							if first_col is None:
								first_col = cell_text
						if len(row_values) >= 10:
							break
					if row_values:
						preview_rows.append(" | ".join(row_values))
					if first_col and len(col1_values) < 8:
						col1_values.append(first_col)
					if len(preview_rows) >= 2:
						break
			summary_parts: list[str] = []
			if names:
				summary_parts.append(f"Sheets: {', '.join(names[:5])}")
			if preview_rows:
				summary_parts.append("Preview: " + " || ".join(preview_rows))
			if col1_values:
				summary_parts.append("Col1: " + " | ".join(col1_values[:8]))
			summary = " ".join(summary_parts).strip()
			return summary[:800] if summary else None
		except Exception:
			return None
